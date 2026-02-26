from __future__ import annotations

import argparse
import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from notion_client import NotionClient

try:
    import openpyxl
except ImportError as e:  # pragma: no cover
    raise SystemExit("openpyxl is required. Install with: python -m pip install openpyxl") from e


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX_PATH = SCRIPT_DIR.parent.parent / "CCKSY_LW_TC_260220.xlsx"
DEFAULT_SHEET_NAME = "???????"

TITLE_PROPERTY = "Scenario ID"
FALLBACK_SHEET_INDEX = 1  # second worksheet in the workbook

# Excel columns that should be mirrored into the Notion DB as-is (rich_text), excluding the primary ID column.
EXCEL_MIRROR_COLUMNS = [
    "1depth",
    "2depth",
    "3depth",
    "?????",
    "?????",
    "???",
    "????",
    "?????(????)",
    "????",
    "????",
    "??(PASS/FAIL/BLOCK)",
    "???(Critical/High/Medium/Low)",
    "??ID",
    "?????/????",
    "????(????/??)",
    "?????(React-Django-Spring)",
    "??????(??/???)",
    "????",
    "???",
    "????(Yes/No)",
    "?????",
    "??????",
    "???",
    "????(Open/In Progress/Resolved/Closed)",
    "??????",
    "??",
]

STATUS_ALLOWED = {"Draft", "Ready", "Running", "PASS", "FAIL", "BLOCK"}
PRIORITY_ALLOWED = {"High", "Medium", "Low"}
SCOPE_ALLOWED = {"Spring API", "React UI", "Integration", "Security"}


def normalize_notion_id(raw: str) -> str:
    value = (raw or "").strip()
    if not value:
        return ""
    compact = re.sub(r"[^0-9a-fA-F]", "", value)
    m = re.search(r"([0-9a-fA-F]{32})", compact)
    if m:
        hex_id = m.group(1).lower()
        return f"{hex_id[:8]}-{hex_id[8:12]}-{hex_id[12:16]}-{hex_id[16:20]}-{hex_id[20:]}"
    return value


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import Excel test scenarios into a Notion test database (upsert by Scenario ID).")
    parser.add_argument("--database-id", default=os.getenv("NOTION_TEST_DB_ID", "").strip(), help="Notion database ID")
    parser.add_argument("--xlsx-path", default=str(DEFAULT_XLSX_PATH), help="Excel workbook path")
    parser.add_argument("--sheet-name", default=DEFAULT_SHEET_NAME, help="Worksheet name (falls back to second sheet if not found)")
    parser.add_argument("--skip-existing", action="store_true", help="Skip existing rows instead of updating")
    return parser.parse_args()


def to_text_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def chunk_text(s: str, size: int = 1800) -> list[str]:
    if not s:
        return []
    return [s[i:i + size] for i in range(0, len(s), size)]


def rich_text_property(value: str) -> dict[str, Any]:
    chunks = chunk_text(value)
    if not chunks:
        return {"rich_text": []}
    return {
        "rich_text": [
            {"type": "text", "text": {"content": part}}
            for part in chunks
        ]
    }


def title_property(value: str) -> dict[str, Any]:
    chunks = chunk_text(value)
    if not chunks:
        chunks = [""]
    return {
        "title": [
            {"type": "text", "text": {"content": part}}
            for part in chunks[:1]
        ]
    }


def select_property(value: str | None) -> dict[str, Any]:
    if not value:
        return {"select": None}
    return {"select": {"name": value}}


def checkbox_property(value: bool) -> dict[str, Any]:
    return {"checkbox": bool(value)}


def date_property_from_text(value: str) -> dict[str, Any]:
    if not value:
        return {"date": None}
    # Keep this conservative: Notion date property receives YYYY-MM-DD only.
    m = re.search(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", value)
    if not m:
        return {"date": None}
    y, mo, d = m.groups()
    try:
        dt = date(int(y), int(mo), int(d))
    except ValueError:
        return {"date": None}
    return {"date": {"start": dt.isoformat()}}


def extract_title_plain(page: dict[str, Any], prop_name: str) -> str:
    props = page.get("properties", {})
    prop = props.get(prop_name, {})
    title_items = prop.get("title", [])
    return "".join(part.get("plain_text", "") for part in title_items).strip()


def load_existing_pages(client: NotionClient, database_id: str) -> dict[str, str]:
    rows: dict[str, str] = {}
    cursor: str | None = None
    while True:
        payload: dict[str, Any] = {}
        if cursor:
            payload["start_cursor"] = cursor
        result = client.query_database(database_id, payload)
        for page in result.get("results", []):
            sid = extract_title_plain(page, TITLE_PROPERTY)
            pid = page.get("id")
            if sid and pid:
                rows[sid] = pid
        if not result.get("has_more"):
            break
        cursor = result.get("next_cursor")
        if not cursor:
            break
    return rows


def read_excel_rows(xlsx_path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, str]]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[FALLBACK_SHEET_INDEX]

    headers = [to_text_value(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
    rows: list[dict[str, str]] = []
    for r in range(2, ws.max_row + 1):
        row_map: dict[str, str] = {}
        has_any = False
        for c, header in enumerate(headers, start=1):
            if not header:
                continue
            raw = ws.cell(row=r, column=c).value
            txt = to_text_value(raw)
            if txt:
                has_any = True
            row_map[header] = txt
        if not has_any:
            continue
        if not row_map.get("ID"):
            continue
        rows.append(row_map)
    return headers, rows


def ensure_excel_mirror_properties(client: NotionClient, database_id: str, db_props: dict[str, Any], headers: list[str]) -> None:
    missing: dict[str, Any] = {}
    for header in headers:
        if not header or header == "ID":
            continue
        if header in db_props:
            continue
        # Mirror raw Excel columns as rich_text to preserve content safely.
        missing[header] = {"rich_text": {}}
    if missing:
        client.update_database(database_id, missing)


def derive_priority(row: dict[str, str]) -> str | None:
    severity = row.get("???(Critical/High/Medium/Low)", "").strip()
    if not severity:
        return None
    if severity == "Critical":
        return "High"
    if severity in PRIORITY_ALLOWED:
        return severity
    return None


def derive_scope(row: dict[str, str]) -> str | None:
    one = row.get("1depth", "").strip()
    two = row.get("2depth", "").strip()
    backend = row.get("?????(React-Django-Spring)", "")
    if "??" in (one + two):
        return "Security"
    if "E2E" in two or "?????" in one:
        return "Integration"
    if "React" in backend:
        return "React UI"
    if "Spring" in backend:
        return "Spring API"
    return None


def build_properties_for_row(row: dict[str, str], db_props: dict[str, Any]) -> dict[str, Any]:
    props: dict[str, Any] = {}
    scenario_id = row.get("ID", "").strip()
    props[TITLE_PROPERTY] = title_property(scenario_id)

    # Fill the pre-created English summary/template columns when present.
    if "Title" in db_props:
        props["Title"] = rich_text_property(row.get("?????", ""))
    if "Priority" in db_props:
        props["Priority"] = select_property(derive_priority(row))
    if "Scope" in db_props:
        props["Scope"] = select_property(derive_scope(row))
    if "Status" in db_props:
        status = row.get("??(PASS/FAIL/BLOCK)", "").strip()
        props["Status"] = select_property(status if status in STATUS_ALLOWED else None)
    if "Owner" in db_props:
        owner = row.get("???", "").strip() or row.get("???", "").strip()
        props["Owner"] = rich_text_property(owner)
    if "Environment" in db_props:
        props["Environment"] = rich_text_property(row.get("?????(React-Django-Spring)", ""))
    if "Preconditions" in db_props:
        props["Preconditions"] = rich_text_property(row.get("????", ""))
    if "Steps" in db_props:
        props["Steps"] = rich_text_property(row.get("?????(????)", ""))
    if "Expected Result" in db_props:
        props["Expected Result"] = rich_text_property(row.get("????", ""))
    if "Actual Result" in db_props:
        props["Actual Result"] = rich_text_property(row.get("????", ""))
    if "Log Summary" in db_props:
        props["Log Summary"] = rich_text_property(row.get("?????/????", ""))
    if "Retest Required" in db_props:
        props["Retest Required"] = checkbox_property(row.get("????(Yes/No)", "").strip().lower() == "yes")
    if "Executed At" in db_props:
        props["Executed At"] = date_property_from_text(row.get("????", ""))

    # Mirror all Excel columns (except ID) into same-named properties, preserving the workbook content.
    for key, value in row.items():
        if not key or key == "ID" or key == TITLE_PROPERTY:
            continue
        if key not in db_props:
            continue
        ptype = next(iter(db_props[key].keys()), None)
        if ptype == "rich_text":
            props[key] = rich_text_property(value)
        elif ptype == "select":
            props[key] = select_property(value or None)
        elif ptype == "checkbox":
            props[key] = checkbox_property(value.strip().lower() in {"y", "yes", "true", "1"})
        elif ptype == "date":
            props[key] = date_property_from_text(value)
        elif ptype == "url":
            props[key] = {"url": value or None}
        elif ptype == "title":
            props[key] = title_property(value)
        else:
            # Fallback to rich_text if property type is unexpected and editable.
            props[key] = rich_text_property(value)
    return props


def main() -> None:
    args = parse_args()
    if not args.database_id:
        raise SystemExit("Missing database id. Use --database-id or set NOTION_TEST_DB_ID.")

    db_id = normalize_notion_id(args.database_id)
    xlsx_path = Path(args.xlsx_path)
    if not xlsx_path.exists():
        raise SystemExit(f"Excel file not found: {xlsx_path}")

    client = NotionClient.from_env()
    db = client.retrieve_database(db_id)
    db_props = db.get("properties", {})

    headers, rows = read_excel_rows(xlsx_path, args.sheet_name)
    ensure_excel_mirror_properties(client, db_id, db_props, headers)
    # Refresh schema after patching.
    db = client.retrieve_database(db_id)
    db_props = db.get("properties", {})

    existing = load_existing_pages(client, db_id)

    created = 0
    updated = 0
    skipped = 0

    for row in rows:
        sid = row.get("ID", "").strip()
        if not sid:
            continue
        props = build_properties_for_row(row, db_props)
        page_id = existing.get(sid)
        if page_id:
            if args.skip_existing:
                skipped += 1
                print(f"- skip existing: {sid}")
                continue
            client.update_page_properties(page_id, props)
            updated += 1
            print(f"- updated: {sid}")
        else:
            client.create_database_page(db_id, props)
            created += 1
            print(f"- created: {sid}")

    print(f"Done. created={created} updated={updated} skipped={skipped} total_excel_rows={len(rows)}")


if __name__ == "__main__":
    main()
